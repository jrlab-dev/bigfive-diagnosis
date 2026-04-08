"""
電子納品PDFファイル名変換スクリプト
- ORGフォルダのM0001_*.PDFをテキスト判定して分かりやすい名前でコピー
- Excel一覧表（納品ファイル一覧表_新.xlsx）を作成
"""

import os
import re
import shutil
from datetime import date

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ===== 設定 =====
ORG_FOLDER = "//N1706wudp/user-o/3 加藤/4.東京都/建設事務所/六建/道路標識設置工事（７六ー１）/1提出書類/電子納品/CD電子納品/PRODUCT/MEET/ORG"
EXCEL_OUT   = "//N1706wudp/user-o/3 加藤/4.東京都/建設事務所/六建/道路標識設置工事（７六ー１）/1提出書類/電子納品/CD電子納品/PRODUCT/納品ファイル一覧表_新.xlsx"
DRY_RUN = True   # True=確認のみ（コピーなし）/ False=実際にコピー＆Excel作成


def safe_name(name: str) -> str:
    """ファイル名に使えない文字をアンダースコアに変換"""
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def detect_name(pdf_path: str) -> str:
    """PDFのテキストを読んで書類名を判定する"""
    try:
        doc = fitz.open(pdf_path)
        raw_text = ""
        for page in doc:
            raw_text += page.get_text()
        doc.close()
    except Exception as e:
        print(f"  [ERROR] {os.path.basename(pdf_path)}: {e}")
        return "読み取りエラー"

    # 行リスト（strip・全角スペース除去）
    lines = [l.strip().replace('\u3000', '').replace('　', '') for l in raw_text.split('\n')]
    lines = [l for l in lines if l]

    # 行ベース判定（「（」→内容→「）記録の報告書」の3行パターン） ※優先度1
    for i in range(len(lines) - 2):
        if lines[i] == '（' and '）記録の報告書' in lines[i + 2]:
            content = lines[i + 1].strip().replace(' ', '').replace('\u3000', '')
            return content + '記録の報告書'
        # 別パターン: 「下記工事の（品質管理）記録の報告書」が分割される場合
        if lines[i].endswith('（') and i + 2 < len(lines) and '）記録の報告書' in lines[i + 2]:
            content = lines[i + 1].strip().replace(' ', '')
            return content + '記録の報告書'

    # 「）記録を報告します」パターン（内容が空白で種別不明の報告書） ※優先度2（上記に引っかからなかった場合のみ）
    for line in lines:
        if '）記録を報告します' in line:
            return '記録の報告書（内容不明）'

    # 全テキストを1行に結合（改行・スペース除去）
    full = raw_text.replace('\n', '').replace('\u3000', '').replace('　', '').replace(' ', '')

    # キーワード判定（優先度順）
    checks = [
        ('品質管理）記録の報告書',     '品質管理記録の報告書'),
        ('安全管理）記録の報告書',     '安全管理記録の報告書'),
        ('出来高数量）記録の報告書',   '出来高数量記録の報告書'),
        ('過積載防止措置）記録の報告書', '過積載防止措置記録の報告書'),
        ('出来形管理）記録の報告書',   '出来形管理記録の報告書'),
        ('工事打合せ）記録の報告書',   '工事打合せ記録の報告書'),
        ('環境物品使用実績',          '環境物品使用実績チェックリスト報告書'),
        ('COBRIS',                   'COBRIS資源利用促進報告書'),
        ('現場閉所報告書',            '現場閉所報告書'),
        ('週休２日制',               '現場閉所報告書'),
        ('完了届',                   '完了届'),
        ('材料確認',                 '材料確認願'),
        ('別記様式甲第136号',        '材料確認願'),
        ('快適トイレ',               '快適トイレ設置協議書'),
        ('建設業退職金共済',         '建設業退職金共済加入届'),
        ('公共事業施行通知書',       '公共事業施行通知書'),
        ('施工計画書',              '施工計画書'),
        ('購入状況報告書',          '購入状況報告書'),
        ('工事履行報告書',          '工事履行報告書'),
        ('段階確認書',             '段階確認書'),
        ('立会確認書',             '立会確認書'),
        ('指示書',                 '指示書'),
        ('協議書',                 '協議書（汎用）'),
    ]

    for keyword, label in checks:
        if keyword in full:
            return label

    # テキストがほぼない場合は白紙
    if len(full) < 20:
        return '白紙'

    return '汎用帳票'


def make_border():
    """細罫線スタイルを返す"""
    side = Side(style='thin')
    return Border(left=side, right=side, top=side, bottom=side)


def main():
    print("=" * 60)
    print("電子納品PDFファイル名変換スクリプト")
    print(f"モード: {'【確認のみ】' if DRY_RUN else '【本番実行】'}")
    print("=" * 60)

    # PDFファイル一覧取得（M0001_*.PDF のみ・ソート）
    all_files = sorted([
        f for f in os.listdir(ORG_FOLDER)
        if re.match(r'M0001_\d+\.PDF', f, re.IGNORECASE)
    ])
    print(f"\n対象ファイル数: {len(all_files)} 件\n")

    # 各ファイルの判定
    results = []  # (元ファイル名, 判定名)
    for fname in all_files:
        fpath = os.path.join(ORG_FOLDER, fname)
        name = detect_name(fpath)
        print(f"  {fname}  →  {name}")
        results.append((fname, name))

    # 重複カウント → 新ファイル名決定
    name_count = {}
    for _, name in results:
        if name != '白紙' and name != '読み取りエラー':
            name_count[name] = name_count.get(name, 0) + 1

    name_used = {}
    rows = []
    for orig_fname, name in results:
        if name in ('白紙', '読み取りエラー'):
            new_fname = name
        else:
            total = name_count.get(name, 1)
            if total == 1:
                new_fname = safe_name(name) + '.PDF'
            else:
                name_used[name] = name_used.get(name, 0) + 1
                new_fname = safe_name(name) + f'_{name_used[name]:02d}.PDF'
        rows.append((orig_fname, name, new_fname))

    # 結果表示
    print("\n" + "-" * 60)
    print("【変換結果一覧】")
    print(f"{'元ファイル名':<20} {'新ファイル名'}")
    print("-" * 60)
    for orig, _, new in rows:
        print(f"{orig:<20} {new}")

    # コピー処理（DRY_RUN=Falseのみ）
    copied = 0
    skipped = 0
    errors = 0
    if not DRY_RUN:
        print("\n【ファイルコピー中...】")
        for orig_fname, name, new_fname in rows:
            if name in ('白紙', '読み取りエラー'):
                print(f"  スキップ: {orig_fname} ({name})")
                skipped += 1
                continue
            src = os.path.join(ORG_FOLDER, orig_fname)
            dst = os.path.join(ORG_FOLDER, new_fname)
            try:
                shutil.copy2(src, dst)
                print(f"  コピー: {orig_fname} → {new_fname}")
                copied += 1
            except Exception as e:
                print(f"  [ERROR] コピー失敗: {orig_fname}: {e}")
                errors += 1

    # Excel作成
    if not DRY_RUN:
        print("\n【Excel一覧表作成中...】")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "電子納品ファイル一覧表"

        # ヘッダー部
        ws['A1'] = '電子納品ファイル一覧表'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = str(date.today())

        # 列ヘッダー（9行目）
        header_row = 9
        headers = ['No.', '保存フォルダ', 'サブフォルダ', '元ファイル名', '新ファイル名', '内容説明']
        hdr_fill = PatternFill(fill_type='solid', fgColor='BDD7EE')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = make_border()

        # 列幅
        col_widths = [8, 10, 10, 25, 40, 40]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

        # データ行
        gray_fill = PatternFill(fill_type='solid', fgColor='EEEEEE')
        for i, (orig_fname, name, new_fname) in enumerate(rows):
            row = header_row + 1 + i
            is_blank = name in ('白紙', '読み取りエラー')

            values = [
                i + 1,
                'MEET',
                'ORG',
                orig_fname,
                '白紙（コピーなし）' if is_blank else new_fname,
                name,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = make_border()
                if is_blank:
                    cell.fill = gray_fill

        wb.save(EXCEL_OUT)
        print(f"  Excel保存完了: {EXCEL_OUT}")

    # サマリー
    total = len(rows)
    blank_count = sum(1 for _, n, _ in rows if n in ('白紙', '読み取りエラー'))
    unknown_count = sum(1 for _, n, _ in rows if n == '汎用帳票')
    print("\n" + "=" * 60)
    print("【処理サマリー】")
    print(f"  総ファイル数: {total}")
    if not DRY_RUN:
        print(f"  コピー完了:   {copied}")
        print(f"  スキップ:     {skipped}（白紙・エラー）")
        print(f"  エラー:       {errors}")
    print(f"  白紙/エラー:  {blank_count}")
    print(f"  判定不明:     {unknown_count}（汎用帳票）")
    print("=" * 60)
    if DRY_RUN:
        print("\n※ DRY_RUN=True のため実際のコピー・Excel作成は行っていません。")
        print("  問題なければ DRY_RUN = False に変更して再実行してください。")


if __name__ == '__main__':
    main()
